"""
Excel satış raporu yazıcı.
4 sayfa: OZET, SIPARISLER, URUN_PERFORMANSI, ULKE_DAGILIMI
"""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from engine.analyzer import (
    calculate_period_metrics,
    get_country_breakdown,
    get_daily_revenue,
    get_top_sellers,
)
from models.order import Order, Platform

# ── Stil Sabitleri ────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2E86AB")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="444444")
NORMAL_FONT = Font(name="Calibri", size=10)
MONEY_FORMAT = '#,##0.00 $'
PERCENT_FORMAT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# KPI kartları için renkler
KPI_FILLS = {
    "green": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "blue": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "orange": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "red": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    "purple": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
}
ETSY_FILL = PatternFill(start_color="FFF0E6", end_color="FFF0E6", fill_type="solid")
AMAZON_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")


def _apply_header_row(ws, row: int, col_start: int, col_end: int):
    """Başlık satırına stil uygular."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _apply_data_row(ws, row: int, col_start: int, col_end: int):
    """Veri satırına stil uygular."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    """Sütun genişliklerini otomatik ayarlar."""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = min(cell_len + 2, max_width)
        ws.column_dimensions[col_letter].width = max_len


def generate_report(
    orders: list[Order],
    products: list,
    output_path: Path,
    period_days: int = 30,
    store_name: str = "Mağaza",
) -> Path:
    """
    Excel satış raporu oluşturur.

    Returns: oluşturulan dosya yolu
    """
    wb = Workbook()

    _write_summary_sheet(wb, orders, products, period_days, store_name)
    _write_orders_sheet(wb, orders)
    _write_product_sheet(wb, orders, products)
    _write_country_sheet(wb, orders)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════
#  SAYFA 1: ÖZET
# ══════════════════════════════════════════════════════════
def _write_summary_sheet(wb, orders, products, period_days, store_name):
    ws = wb.active
    ws.title = "OZET"
    ws.sheet_properties.tabColor = "2E86AB"

    today = date.today()
    period_start = today - timedelta(days=period_days)
    prev_start = period_start - timedelta(days=period_days)
    prev_end = period_start - timedelta(days=1)

    current = calculate_period_metrics(orders, period_start, today)
    previous = calculate_period_metrics(orders, prev_start, prev_end)

    # Başlık
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{store_name} - Satış Raporu"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Rapor Tarihi: {today.strftime('%d.%m.%Y')} | Dönem: Son {period_days} gün"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── KPI Kartları ──
    row = 4
    kpis = [
        ("Toplam Sipariş", current.total_orders, previous.total_orders, "green", None),
        ("Brüt Gelir", current.gross_revenue, previous.gross_revenue, "blue", MONEY_FORMAT),
        ("Net Gelir", current.net_revenue, previous.net_revenue, "orange", MONEY_FORMAT),
        ("Ort. Sipariş Değeri", current.avg_order_value, previous.avg_order_value, "purple", MONEY_FORMAT),
        ("Satılan Ürün", current.total_items_sold, previous.total_items_sold, "green", None),
        ("Benzersiz Müşteri", current.unique_buyers, previous.unique_buyers, "blue", None),
    ]

    headers_kpi = ["Metrik", "Bu Dönem", "Önceki Dönem", "Değişim"]
    for i, h in enumerate(headers_kpi, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, 1, 4)

    for metric_name, curr_val, prev_val, color, fmt in kpis:
        row += 1
        ws.cell(row=row, column=1, value=metric_name)
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10)

        cell_curr = ws.cell(row=row, column=2, value=curr_val)
        cell_prev = ws.cell(row=row, column=3, value=prev_val)
        if fmt:
            cell_curr.number_format = fmt
            cell_prev.number_format = fmt

        if prev_val and prev_val > 0:
            change = (curr_val - prev_val) / prev_val
            cell_change = ws.cell(row=row, column=4, value=change)
            cell_change.number_format = PERCENT_FORMAT
            if change > 0:
                cell_change.font = Font(name="Calibri", color="2E7D32", bold=True)
            elif change < 0:
                cell_change.font = Font(name="Calibri", color="C62828", bold=True)
        else:
            ws.cell(row=row, column=4, value="-")

        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = KPI_FILLS[color]
            ws.cell(row=row, column=col).border = THIN_BORDER

    # ── Platform Kırılımı ──
    row += 2
    ws.cell(row=row, column=1, value="Platform Kırılımı")
    ws.cell(row=row, column=1).font = SUBTITLE_FONT

    row += 1
    platform_headers = ["Platform", "Sipariş", "Brüt Gelir", "Net Gelir", "Ort. Sipariş"]
    for i, h in enumerate(platform_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, 1, 5)

    for platform, fill in [(Platform.ETSY, ETSY_FILL), (Platform.AMAZON, AMAZON_FILL)]:
        p_orders = [o for o in orders if o.platform == platform]
        if not p_orders:
            continue
        row += 1
        p_metrics = calculate_period_metrics(p_orders, period_start, today)

        ws.cell(row=row, column=1, value=platform.value.upper())
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True)
        ws.cell(row=row, column=2, value=p_metrics.total_orders)
        ws.cell(row=row, column=3, value=p_metrics.gross_revenue)
        ws.cell(row=row, column=3).number_format = MONEY_FORMAT
        ws.cell(row=row, column=4, value=p_metrics.net_revenue)
        ws.cell(row=row, column=4).number_format = MONEY_FORMAT
        ws.cell(row=row, column=5, value=p_metrics.avg_order_value)
        ws.cell(row=row, column=5).number_format = MONEY_FORMAT

        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = THIN_BORDER

    # ── En Çok Satanlar ──
    row += 2
    ws.cell(row=row, column=1, value="En Çok Satan 5 Ürün")
    ws.cell(row=row, column=1).font = SUBTITLE_FONT

    row += 1
    top_headers = ["#", "Ürün", "Adet", "Gelir"]
    for i, h in enumerate(top_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_row(ws, row, 1, 4)

    top_sellers = get_top_sellers(orders, limit=5)
    for rank, ts in enumerate(top_sellers, 1):
        row += 1
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=ts.title[:50])
        ws.cell(row=row, column=3, value=ts.units_sold)
        ws.cell(row=row, column=4, value=ts.revenue)
        ws.cell(row=row, column=4).number_format = MONEY_FORMAT
        _apply_data_row(ws, row, 1, 4)

    # ── Günlük Gelir Grafiği ──
    row += 2
    ws.cell(row=row, column=1, value="Günlük Gelir (Son 30 Gün)")
    ws.cell(row=row, column=1).font = SUBTITLE_FONT

    daily = get_daily_revenue(orders, days=30)
    chart_start_row = row + 1
    row += 1
    ws.cell(row=row, column=1, value="Tarih")
    ws.cell(row=row, column=2, value="Gelir ($)")
    _apply_header_row(ws, row, 1, 2)

    for d, rev in sorted(daily.items()):
        row += 1
        ws.cell(row=row, column=1, value=d.strftime("%d.%m"))
        ws.cell(row=row, column=2, value=rev)
        ws.cell(row=row, column=2).number_format = MONEY_FORMAT
        _apply_data_row(ws, row, 1, 2)

    chart = LineChart()
    chart.title = "Günlük Gelir Trendi"
    chart.style = 10
    chart.y_axis.title = "Gelir ($)"
    chart.x_axis.title = "Tarih"
    chart.width = 25
    chart.height = 12

    data_ref = Reference(ws, min_col=2, min_row=chart_start_row, max_row=row)
    cats_ref = Reference(ws, min_col=1, min_row=chart_start_row + 1, max_row=row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.width = 25000

    ws.add_chart(chart, f"D{chart_start_row}")

    _auto_width(ws)


# ══════════════════════════════════════════════════════════
#  SAYFA 2: SİPARİŞLER
# ══════════════════════════════════════════════════════════
def _write_orders_sheet(wb, orders):
    ws = wb.create_sheet("SIPARISLER")
    ws.sheet_properties.tabColor = "4CAF50"

    headers = [
        "Tarih", "Platform", "Sipariş No", "Müşteri", "Ülke",
        "Ürünler", "Adet", "Brüt Gelir", "Kargo", "Vergi",
        "İndirim", "Platform Kesintisi", "Net Gelir", "Durum",
    ]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _apply_header_row(ws, 1, 1, len(headers))

    sorted_orders = sorted(orders, key=lambda o: o.order_date, reverse=True)

    for row_idx, o in enumerate(sorted_orders, 2):
        items_str = ", ".join(item.product_title[:30] for item in o.items[:3])
        if len(o.items) > 3:
            items_str += f" +{len(o.items) - 3} daha"

        values = [
            o.order_date.strftime("%d.%m.%Y %H:%M"),
            o.platform.value.upper(),
            o.order_id,
            o.buyer_name or "-",
            o.buyer_country or "-",
            items_str,
            o.item_count,
            o.gross_revenue,
            o.shipping_cost,
            o.tax,
            o.discount,
            o.total_fees,
            o.net_revenue,
            o.status.value,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

        # Para formatı
        for col in [8, 9, 10, 11, 12, 13]:
            ws.cell(row=row_idx, column=col).number_format = MONEY_FORMAT

        # Platform rengi
        fill = ETSY_FILL if o.platform == Platform.ETSY else AMAZON_FILL
        ws.cell(row=row_idx, column=2).fill = fill

    # Toplam satırı
    total_row = len(sorted_orders) + 2
    ws.cell(row=total_row, column=1, value="TOPLAM")
    ws.cell(row=total_row, column=1).font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=total_row, column=7, value=sum(o.item_count for o in orders))
    ws.cell(row=total_row, column=8, value=sum(o.gross_revenue for o in orders))
    ws.cell(row=total_row, column=8).number_format = MONEY_FORMAT
    ws.cell(row=total_row, column=13, value=sum(o.net_revenue for o in orders))
    ws.cell(row=total_row, column=13).number_format = MONEY_FORMAT

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(name="Calibri", bold=True)
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{total_row - 1}"
    ws.freeze_panes = "A2"
    _auto_width(ws)


# ══════════════════════════════════════════════════════════
#  SAYFA 3: ÜRÜN PERFORMANSI
# ══════════════════════════════════════════════════════════
def _write_product_sheet(wb, orders, products):
    ws = wb.create_sheet("URUN_PERFORMANSI")
    ws.sheet_properties.tabColor = "FF9800"

    headers = [
        "Platform", "Ürün", "Fiyat", "Stok", "Görüntülenme",
        "Favori", "Toplam Satış", "Toplam Gelir", "Dönüşüm %",
        "Favori Oranı %", "Durum", "Uyarı",
    ]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _apply_header_row(ws, 1, 1, len(headers))

    # Sipariş verisinden satış hesapla
    sales_by_product = {}
    for o in orders:
        for item in o.items:
            pid = item.product_id
            if pid not in sales_by_product:
                sales_by_product[pid] = {"units": 0, "revenue": 0.0}
            sales_by_product[pid]["units"] += item.quantity
            sales_by_product[pid]["revenue"] += item.total_price

    for row_idx, p in enumerate(products, 2):
        sale_info = sales_by_product.get(p.product_id, {"units": 0, "revenue": 0.0})

        # Uyarı belirleme
        alert = ""
        if p.quantity == 0:
            alert = "STOK BİTTİ!"
        elif p.quantity <= 5:
            alert = "DÜŞÜK STOK"
        elif p.views > 100 and p.conversion_rate < 1.0:
            alert = "DÜŞÜK DÖNÜŞÜM"
        elif p.favorites > 20 and sale_info["units"] < 3:
            alert = "FAVORİ AMA SATMIYOR"

        values = [
            p.platform.value.upper(),
            p.title[:50],
            p.price,
            p.quantity,
            p.views,
            p.favorites,
            sale_info["units"],
            sale_info["revenue"],
            p.conversion_rate / 100 if p.conversion_rate else 0,
            p.favorite_rate / 100 if p.favorite_rate else 0,
            p.status,
            alert,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

        ws.cell(row=row_idx, column=3).number_format = MONEY_FORMAT
        ws.cell(row=row_idx, column=8).number_format = MONEY_FORMAT
        ws.cell(row=row_idx, column=9).number_format = PERCENT_FORMAT
        ws.cell(row=row_idx, column=10).number_format = PERCENT_FORMAT

        # Platform rengi
        fill = ETSY_FILL if p.platform == Platform.ETSY else AMAZON_FILL
        ws.cell(row=row_idx, column=1).fill = fill

        # Uyarı rengi
        if alert:
            alert_cell = ws.cell(row=row_idx, column=12)
            if "BİTTİ" in alert:
                alert_cell.fill = ALERT_FILL
                alert_cell.font = Font(name="Calibri", bold=True, color="C62828")
            else:
                alert_cell.fill = WARNING_FILL
                alert_cell.font = Font(name="Calibri", bold=True, color="E65100")

    # Grafik - En çok satanlar
    top = get_top_sellers(orders, limit=8)
    if top:
        chart_row = len(products) + 3
        ws.cell(row=chart_row, column=1, value="Ürün")
        ws.cell(row=chart_row, column=2, value="Gelir ($)")
        ws.cell(row=chart_row, column=3, value="Adet")

        for i, ts in enumerate(top):
            r = chart_row + 1 + i
            ws.cell(row=r, column=1, value=ts.title[:25])
            ws.cell(row=r, column=2, value=ts.revenue)
            ws.cell(row=r, column=3, value=ts.units_sold)

        chart = BarChart()
        chart.type = "col"
        chart.title = "En Çok Satan Ürünler"
        chart.y_axis.title = "Gelir ($)"
        chart.width = 25
        chart.height = 14

        data_ref = Reference(ws, min_col=2, max_col=2, min_row=chart_row, max_row=chart_row + len(top))
        cats_ref = Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + len(top))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4

        ws.add_chart(chart, f"E{chart_row}")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(products) + 1}"
    ws.freeze_panes = "A2"
    _auto_width(ws)


# ══════════════════════════════════════════════════════════
#  SAYFA 4: ÜLKE DAĞILIMI
# ══════════════════════════════════════════════════════════
def _write_country_sheet(wb, orders):
    ws = wb.create_sheet("ULKE_DAGILIMI")
    ws.sheet_properties.tabColor = "9C27B0"

    headers = ["Ülke", "Sipariş Sayısı", "Toplam Gelir", "Ort. Sipariş", "Pay %"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _apply_header_row(ws, 1, 1, len(headers))

    countries = get_country_breakdown(orders)
    total_orders = len(orders)
    total_revenue = sum(o.gross_revenue for o in orders)

    row = 2
    for country, count in countries.items():
        country_orders = [o for o in orders if o.buyer_country == country]
        country_revenue = sum(o.gross_revenue for o in country_orders)
        avg_order = country_revenue / count if count > 0 else 0
        share = count / total_orders if total_orders > 0 else 0

        ws.cell(row=row, column=1, value=country)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=country_revenue)
        ws.cell(row=row, column=3).number_format = MONEY_FORMAT
        ws.cell(row=row, column=4, value=avg_order)
        ws.cell(row=row, column=4).number_format = MONEY_FORMAT
        ws.cell(row=row, column=5, value=share)
        ws.cell(row=row, column=5).number_format = PERCENT_FORMAT
        _apply_data_row(ws, row, 1, 5)
        row += 1

    # Toplam
    ws.cell(row=row, column=1, value="TOPLAM")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True)
    ws.cell(row=row, column=2, value=total_orders)
    ws.cell(row=row, column=3, value=total_revenue)
    ws.cell(row=row, column=3).number_format = MONEY_FORMAT
    ws.cell(row=row, column=5, value=1.0)
    ws.cell(row=row, column=5).number_format = PERCENT_FORMAT
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", bold=True)
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Pasta grafik
    if len(countries) > 1:
        chart = PieChart()
        chart.title = "Ülke Dağılımı"
        chart.width = 18
        chart.height = 14

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=min(row - 1, 11))
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=min(row - 1, 11))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False

        ws.add_chart(chart, "G2")

    _auto_width(ws)
